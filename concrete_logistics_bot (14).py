"""
Concrete Logistics Telegram Bot
================================
Two access modes:
  - Admin link:  t.me/YourBot?start=admin_SECRET123
  - Worker link: t.me/YourBot?start=worker_SECRET456

Features:
- Separate Admin / Worker access via deep links
- Truck plate: pick from saved list OR add manually
- Reminder after truck plate: confirm details before saving
- Job Management: Add, Status, Cancel, Complete (admin only)
- Daily / Weekly / Monthly Reports + Excel Export
- SQLite storage

Requirements:
    pip install python-telegram-bot openpyxl

Setup:
    1. Set BOT_TOKEN below
    2. Set ADMIN_SECRET and WORKER_SECRET (any password you choose)
    3. Run: python3 concrete_logistics_bot.py
    4. Share links:
       Admin:  https://t.me/YourBotUsername?start=admin_SECRET123
       Worker: https://t.me/YourBotUsername?start=worker_SECRET456
"""

import sqlite3
import logging
import io
from datetime import datetime, timedelta, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputFile,
)
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ContextTypes,
    ConversationHandler,
    filters,
)

# ─────────────────────────────────────────────
# CONFIG  ← Edit these!
# ─────────────────────────────────────────────
BOT_TOKEN     = "8468077984:AAGGaDOML_5oA1DGz3rYCsUjmkOX8Xxj2kM"
BOT_USERNAME  = "YourBotUsername"   # ← Your bot's @username (without @)

ADMIN_SECRET  = "admin_SECRET123"   # ← Change this to something secret
WORKER_SECRET = "worker_SECRET456"  # ← Change this to something secret

DB_PATH = "logistics.db"

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# Conversation states
ASK_JOB, ASK_PLATE, ASK_PLATE_MANUAL, ASK_VOLUME, CONFIRM_TRIP = range(5)
ASK_NEW_JOB_NAME, ASK_NEW_JOB_LOCATION = range(10, 12)
ASK_NEW_TRUCK = 20

STATUS_EMOJI = {"active": "🟢", "completed": "✅", "cancelled": "❌"}


# ─────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────
def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                user_id   INTEGER PRIMARY KEY,
                username  TEXT,
                role      TEXT NOT NULL DEFAULT 'worker',
                joined_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS trips (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id     INTEGER NOT NULL,
                timestamp   TEXT    NOT NULL,
                job_name    TEXT    NOT NULL,
                truck_plate TEXT    NOT NULL,
                volume      REAL    NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS jobs (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT NOT NULL,
                location   TEXT,
                status     TEXT NOT NULL DEFAULT 'active',
                created_at TEXT NOT NULL,
                updated_at TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS trucks (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                plate      TEXT NOT NULL UNIQUE,
                added_at   TEXT NOT NULL
            )
        """)
        conn.commit()


# ── Users ─────────────────────────────────────
def register_user(user_id: int, username: str, role: str):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            INSERT INTO users (user_id, username, role, joined_at)
            VALUES (?,?,?,?)
            ON CONFLICT(user_id) DO UPDATE SET role=excluded.role, username=excluded.username
        """, (user_id, username or "", role, datetime.now().isoformat(sep=" ", timespec="seconds")))
        conn.commit()


def get_user_role(user_id: int) -> str:
    with sqlite3.connect(DB_PATH) as conn:
        row = conn.execute("SELECT role FROM users WHERE user_id=?", (user_id,)).fetchone()
    return row[0] if row else None


def is_admin(user_id: int) -> bool:
    return get_user_role(user_id) == "admin"


# ── Trucks ────────────────────────────────────
def get_all_trucks() -> list:
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM trucks ORDER BY plate").fetchall()
    return [dict(r) for r in rows]


def add_truck(plate: str):
    with sqlite3.connect(DB_PATH) as conn:
        try:
            conn.execute("INSERT INTO trucks (plate, added_at) VALUES (?,?)",
                         (plate.upper().strip(), datetime.now().isoformat(sep=" ", timespec="seconds")))
            conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False  # already exists


# ── Jobs ──────────────────────────────────────
def get_all_jobs(status_filter=None) -> list:
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        if status_filter:
            rows = conn.execute("SELECT * FROM jobs WHERE status=? ORDER BY created_at DESC", (status_filter,)).fetchall()
        else:
            rows = conn.execute("SELECT * FROM jobs ORDER BY created_at DESC").fetchall()
    return [dict(r) for r in rows]


def get_job_by_id(job_id: int):
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
    return dict(row) if row else None


def add_job(name: str, location: str):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("INSERT INTO jobs (name, location, status, created_at) VALUES (?,?,?,?)",
                     (name, location, "active", datetime.now().isoformat(sep=" ", timespec="seconds")))
        conn.commit()


def update_job_status(job_id: int, status: str):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("UPDATE jobs SET status=?, updated_at=? WHERE id=?",
                     (status, datetime.now().isoformat(sep=" ", timespec="seconds"), job_id))
        conn.commit()


# ── Trips ─────────────────────────────────────
def save_trip(user_id: int, job_name: str, truck_plate: str, volume: float):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            "INSERT INTO trips (user_id, timestamp, job_name, truck_plate, volume) VALUES (?,?,?,?,?)",
            (user_id, datetime.now().isoformat(sep=" ", timespec="seconds"), job_name, truck_plate, volume),
        )
        conn.commit()


def fetch_trips(period: str) -> list:
    today = date.today()
    if period == "daily":
        query, params = "SELECT * FROM trips WHERE DATE(timestamp) = ?", (today.isoformat(),)
    elif period == "weekly":
        monday = today - timedelta(days=today.weekday())
        query, params = "SELECT * FROM trips WHERE DATE(timestamp) >= ?", (monday.isoformat(),)
    else:
        first = today.replace(day=1)
        query, params = "SELECT * FROM trips WHERE DATE(timestamp) >= ?", (first.isoformat(),)
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(query, params).fetchall()
    return [dict(r) for r in rows]


def get_job_trip_summary(job_name: str):
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT COUNT(*) as trips, SUM(volume) as total_volume FROM trips WHERE job_name=?",
            (job_name,)
        ).fetchone()
    return dict(row) if row else {"trips": 0, "total_volume": 0}


# ─────────────────────────────────────────────
# REPORT GENERATION
# ─────────────────────────────────────────────
def _period_label(period: str) -> str:
    today = date.today()
    if period == "daily":
        return f"Daily Report — {today.strftime('%B %d, %Y')}"
    elif period == "weekly":
        monday = today - timedelta(days=today.weekday())
        sunday = monday + timedelta(days=6)
        return f"Weekly Report ({monday.strftime('%b %d')}–{sunday.strftime('%b %d, %Y')})"
    else:
        return f"Monthly Report — {today.strftime('%B %Y')}"


def generate_text_report(period: str) -> str:
    trips = fetch_trips(period)
    label = _period_label(period)
    if not trips:
        return f"📋 *{label}*\n\n_No trips recorded for this period._"

    job_totals: dict = {}
    for t in trips:
        j = t["job_name"]
        if j not in job_totals:
            job_totals[j] = {"volume": 0.0, "trips": 0}
        job_totals[j]["volume"] += t["volume"]
        job_totals[j]["trips"]  += 1

    sorted_jobs = sorted(job_totals.items(), key=lambda x: x[1]["volume"], reverse=True)[:5]
    total_vol   = sum(v["volume"] for _, v in job_totals.items())
    total_trips = sum(v["trips"]  for _, v in job_totals.items())
    job_lines   = "\n".join(
        f"  {i+1}. {name}: *{data['volume']:.1f} m³* ({data['trips']} trips)"
        for i, (name, data) in enumerate(sorted_jobs)
    )

    truck_totals: dict = {}
    for t in trips:
        p = t["truck_plate"]
        if p not in truck_totals:
            truck_totals[p] = {"volume": 0.0, "trips": 0}
        truck_totals[p]["volume"] += t["volume"]
        truck_totals[p]["trips"]  += 1

    sorted_trucks = sorted(truck_totals.items(), key=lambda x: x[1]["volume"], reverse=True)
    truck_lines   = "\n".join(
        f"  {'🥇' if i==0 else '🥈' if i==1 else '🥉' if i==2 else '▪️'} "
        f"{plate}: *{data['volume']:.1f} m³* ({data['trips']} trips)"
        for i, (plate, data) in enumerate(sorted_trucks)
    )

    return (
        f"📋 *{label}*\n━━━━━━━━━━━━━━━━━━━━\n\n"
        f"📅 *Job Site Summary* (Top 5)\n{job_lines}\n"
        f"  ─────────────────\n"
        f"  Total: *{total_vol:.1f} m³* | *{total_trips} trips*\n\n"
        f"🚛 *Truck Performance*\n{truck_lines}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"_Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}_"
    )


def generate_excel_report(period: str) -> io.BytesIO:
    trips  = fetch_trips(period)
    wb     = openpyxl.Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw Trips"
    headers = ["#", "Timestamp", "Job Site", "Truck Plate", "Volume (m³)"]
    hfill   = PatternFill("solid", fgColor="1F4E79")
    for col, h in enumerate(headers, 1):
        cell = ws_raw.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
    for i, t in enumerate(trips, 2):
        ws_raw.cell(row=i, column=1, value=i - 1)
        ws_raw.cell(row=i, column=2, value=t["timestamp"])
        ws_raw.cell(row=i, column=3, value=t["job_name"])
        ws_raw.cell(row=i, column=4, value=t["truck_plate"])
        ws_raw.cell(row=i, column=5, value=t["volume"])
    for col in ws_raw.columns:
        ws_raw.column_dimensions[col[0].column_letter].width = 20

    ws_job = wb.create_sheet("Job Summary")
    ws_job.append(["Job Site", "Total Volume (m³)", "Total Trips"])
    for cell in ws_job[1]: cell.font = Font(bold=True)
    job_totals: dict = {}
    for t in trips:
        j = t["job_name"]
        if j not in job_totals:
            job_totals[j] = {"volume": 0.0, "trips": 0}
        job_totals[j]["volume"] += t["volume"]
        job_totals[j]["trips"]  += 1
    for name, data in sorted(job_totals.items(), key=lambda x: x[1]["volume"], reverse=True):
        ws_job.append([name, round(data["volume"], 2), data["trips"]])

    ws_truck = wb.create_sheet("Truck Rankings")
    ws_truck.append(["Rank", "Truck Plate", "Total Volume (m³)", "Total Trips"])
    for cell in ws_truck[1]: cell.font = Font(bold=True)
    truck_totals: dict = {}
    for t in trips:
        p = t["truck_plate"]
        if p not in truck_totals:
            truck_totals[p] = {"volume": 0.0, "trips": 0}
        truck_totals[p]["volume"] += t["volume"]
        truck_totals[p]["trips"]  += 1
    for rank, (plate, data) in enumerate(
        sorted(truck_totals.items(), key=lambda x: x[1]["volume"], reverse=True), 1
    ):
        ws_truck.append([rank, plate, round(data["volume"], 2), data["trips"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def build_main_menu(user_id: int) -> InlineKeyboardMarkup:
    admin    = is_admin(user_id)
    keyboard = []
    if admin:
        keyboard.append([InlineKeyboardButton("🆕 Add New Job",     callback_data="add_job")])
        keyboard.append([InlineKeyboardButton("🚛 Manage Trucks",   callback_data="manage_trucks")])
    keyboard.append([InlineKeyboardButton("➕ Log New Trip",     callback_data="log_trip")])
    keyboard.append([InlineKeyboardButton("🏗️ Job Status",       callback_data="job_status")])
    keyboard.append([InlineKeyboardButton("📊 View Reports",     callback_data="menu_reports")])
    keyboard.append([InlineKeyboardButton("📥 Export to Excel",  callback_data="menu_export")])
    if admin:
        keyboard.append([InlineKeyboardButton("✅ Complete Job",    callback_data="complete_job")])
        keyboard.append([InlineKeyboardButton("❌ Cancel Job",      callback_data="cancel_job")])
    return InlineKeyboardMarkup(keyboard)


def job_select_keyboard(jobs: list, prefix: str) -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton(
            f"{STATUS_EMOJI.get(j['status'], '⬜')} {j['name']} — {j['location'] or 'No location'}",
            callback_data=f"{prefix}{j['id']}"
        )]
        for j in jobs
    ]
    buttons.append([InlineKeyboardButton("⬅️ Back", callback_data="back_main")])
    return InlineKeyboardMarkup(buttons)


# ─────────────────────────────────────────────
# /start — Deep Link Handler
# ─────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user    = update.effective_user
    user_id = user.id
    args    = context.args  # text after /start

    existing_role = get_user_role(user_id)

    # First time or using a link
    if args:
        token = args[0]
        if token == ADMIN_SECRET:
            register_user(user_id, user.username, "admin")
            await update.message.reply_text(
                f"👑 *Welcome, Admin {user.first_name}!*\n\n"
                f"You have full access to the Concrete Logistics Bot.\n\n"
                f"🔗 *Share these links with your team:*\n"
                f"Admin:  `https://t.me/{BOT_USERNAME}?start={ADMIN_SECRET}`\n"
                f"Worker: `https://t.me/{BOT_USERNAME}?start={WORKER_SECRET}`",
                parse_mode="Markdown",
                reply_markup=build_main_menu(user_id),
            )
            return
        elif token == WORKER_SECRET:
            register_user(user_id, user.username, "worker")
            await update.message.reply_text(
                f"👷 *Welcome, {user.first_name}!*\n\n"
                f"You have worker access to the Concrete Logistics Bot.",
                parse_mode="Markdown",
                reply_markup=build_main_menu(user_id),
            )
            return
        else:
            await update.message.reply_text(
                "❌ *Invalid access link.*\n\nPlease ask your admin for the correct link.",
                parse_mode="Markdown",
            )
            return

    # Returning user (no token needed)
    if existing_role:
        role_label = "👑 Admin" if existing_role == "admin" else "👷 Worker"
        await update.message.reply_text(
            f"🏗️ *Concrete Logistics Bot*\n_{role_label}_\n\nWelcome back, {user.first_name}!",
            parse_mode="Markdown",
            reply_markup=build_main_menu(user_id),
        )
        return

    # Unknown user, no token
    await update.message.reply_text(
        "🚫 *Access Required*\n\n"
        "Please use the link provided by your admin to access this bot.\n\n"
        "Contact your admin to get your access link.",
        parse_mode="Markdown",
    )


async def back_main(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    role    = get_user_role(user_id)
    if not role:
        await query.edit_message_text("🚫 Access denied. Use your access link.")
        return
    role_label = "👑 Admin" if role == "admin" else "👷 Worker"
    await query.edit_message_text(
        f"🏗️ *Concrete Logistics Bot*\n_{role_label}_\n\nWhat would you like to do?",
        parse_mode="Markdown",
        reply_markup=build_main_menu(user_id),
    )


def access_denied_keyboard():
    return InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="back_main")]])


# ─────────────────────────────────────────────
# JOB STATUS (everyone)
# ─────────────────────────────────────────────
async def job_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    user_id = query.from_user.id
    await query.answer()
    if not get_user_role(user_id):
        await query.edit_message_text("🚫 Access denied.", reply_markup=access_denied_keyboard())
        return

    jobs = get_all_jobs()
    if not jobs:
        await query.edit_message_text(
            "🏗️ *Job Status*\n\n_No jobs found. Ask an admin to add jobs._",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="back_main")]]),
        )
        return

    lines = []
    for j in jobs:
        emoji   = STATUS_EMOJI.get(j["status"], "⬜")
        summary = get_job_trip_summary(j["name"])
        vol     = summary["total_volume"] or 0
        trips   = summary["trips"] or 0
        lines.append(
            f"{emoji} *{j['name']}*\n"
            f"   📍 {j['location'] or 'No location'}\n"
            f"   🧱 {vol:.1f} m³ | {trips} trips\n"
            f"   Status: `{j['status'].upper()}`"
        )

    text = "🏗️ *All Job Sites*\n━━━━━━━━━━━━━━━━━━━━\n\n" + "\n\n".join(lines)
    await query.edit_message_text(
        text, parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="back_main")]]),
    )


# ─────────────────────────────────────────────
# MANAGE TRUCKS (admin)
# ─────────────────────────────────────────────
async def manage_trucks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    user_id = query.from_user.id
    await query.answer()

    if not is_admin(user_id):
        await query.edit_message_text("🚫 *Access Denied*", parse_mode="Markdown",
                                       reply_markup=access_denied_keyboard())
        return

    trucks = get_all_trucks()
    if trucks:
        truck_list = "\n".join(f"  🚛 `{t['plate']}`" for t in trucks)
        text = f"🚛 *Registered Trucks*\n━━━━━━━━━━━━━━━━━━━━\n\n{truck_list}\n\n_Tap below to add a new truck._"
    else:
        text = "🚛 *Registered Trucks*\n\n_No trucks registered yet._"

    keyboard = [
        [InlineKeyboardButton("➕ Add New Truck", callback_data="add_truck")],
        [InlineKeyboardButton("⬅️ Back",          callback_data="back_main")],
    ]
    await query.edit_message_text(text, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(keyboard))


async def add_truck_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query   = update.callback_query
    user_id = query.from_user.id
    await query.answer()

    if not is_admin(user_id):
        await query.edit_message_text("🚫 *Access Denied*", parse_mode="Markdown",
                                       reply_markup=access_denied_keyboard())
        return ConversationHandler.END

    await query.message.reply_text("🚛 *Add New Truck*\n\nEnter the truck plate number:", parse_mode="Markdown")
    return ASK_NEW_TRUCK


async def save_new_truck(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    plate = update.message.text.strip().upper()
    added = add_truck(plate)
    if added:
        keyboard = [
            [InlineKeyboardButton("➕ Add Another Truck", callback_data="add_truck")],
            [InlineKeyboardButton("🏠 Main Menu",          callback_data="back_main")],
        ]
        await update.message.reply_text(
            f"✅ *Truck Added!*\n\n🚛 Plate: `{plate}`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
    else:
        await update.message.reply_text(
            f"⚠️ Truck `{plate}` already exists.",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="back_main")]]),
        )
    return ConversationHandler.END


# ─────────────────────────────────────────────
# ADD NEW JOB (admin only)
# ─────────────────────────────────────────────
async def add_job_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query   = update.callback_query
    user_id = query.from_user.id
    await query.answer()

    if not is_admin(user_id):
        await query.edit_message_text("🚫 *Access Denied*\n\nOnly admins can add new jobs.",
                                       parse_mode="Markdown", reply_markup=access_denied_keyboard())
        return ConversationHandler.END

    await query.message.reply_text("🆕 *Add New Job*\n\nEnter the *Job Site Name:*", parse_mode="Markdown")
    return ASK_NEW_JOB_NAME


async def ask_new_job_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["new_job_name"] = update.message.text.strip()
    await update.message.reply_text("📍 Enter the *Job Location* (e.g. Downtown, Block 5):", parse_mode="Markdown")
    return ASK_NEW_JOB_LOCATION


async def ask_new_job_location(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    name     = context.user_data["new_job_name"]
    location = update.message.text.strip()
    add_job(name, location)
    keyboard = [
        [InlineKeyboardButton("🆕 Add Another Job", callback_data="add_job")],
        [InlineKeyboardButton("🏠 Main Menu",        callback_data="back_main")],
    ]
    await update.message.reply_text(
        f"✅ *Job Added!*\n\n🏗️ Name: `{name}`\n📍 Location: `{location}`\n🟢 Status: `ACTIVE`",
        parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(keyboard),
    )
    context.user_data.clear()
    return ConversationHandler.END


# ─────────────────────────────────────────────
# COMPLETE JOB (admin only)
# ─────────────────────────────────────────────
async def complete_job_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    user_id = query.from_user.id
    await query.answer()
    if not is_admin(user_id):
        await query.edit_message_text("🚫 *Access Denied*", parse_mode="Markdown",
                                       reply_markup=access_denied_keyboard())
        return
    jobs = get_all_jobs(status_filter="active")
    if not jobs:
        await query.edit_message_text("✅ *Complete Job*\n\n_No active jobs._", parse_mode="Markdown",
                                       reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="back_main")]]))
        return
    await query.edit_message_text("✅ *Select job to mark COMPLETE:*", parse_mode="Markdown",
                                   reply_markup=job_select_keyboard(jobs, "do_complete_"))


async def do_complete_job(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query  = update.callback_query
    job_id = int(query.data.replace("do_complete_", ""))
    await query.answer()
    job = get_job_by_id(job_id)
    if job:
        update_job_status(job_id, "completed")
        summary = get_job_trip_summary(job["name"])
        vol     = summary["total_volume"] or 0
        trips   = summary["trips"] or 0
        await query.edit_message_text(
            f"✅ *Job Completed!*\n\n🏗️ *{job['name']}*\n📍 {job['location'] or 'No location'}\n"
            f"🧱 Total: *{vol:.1f} m³* ({trips} trips)",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Main Menu", callback_data="back_main")]]),
        )


# ─────────────────────────────────────────────
# CANCEL JOB (admin only)
# ─────────────────────────────────────────────
async def cancel_job_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    user_id = query.from_user.id
    await query.answer()
    if not is_admin(user_id):
        await query.edit_message_text("🚫 *Access Denied*", parse_mode="Markdown",
                                       reply_markup=access_denied_keyboard())
        return
    jobs = get_all_jobs(status_filter="active")
    if not jobs:
        await query.edit_message_text("❌ *Cancel Job*\n\n_No active jobs._", parse_mode="Markdown",
                                       reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="back_main")]]))
        return
    await query.edit_message_text("❌ *Select job to CANCEL:*", parse_mode="Markdown",
                                   reply_markup=job_select_keyboard(jobs, "do_cancel_"))


async def do_cancel_job(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query  = update.callback_query
    job_id = int(query.data.replace("do_cancel_", ""))
    await query.answer()
    job = get_job_by_id(job_id)
    if job:
        update_job_status(job_id, "cancelled")
        await query.edit_message_text(
            f"❌ *Job Cancelled*\n\n🏗️ *{job['name']}*\n📍 {job['location'] or 'No location'}\nStatus: `CANCELLED`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Main Menu", callback_data="back_main")]]),
        )


# ─────────────────────────────────────────────
# REPORT / EXPORT MENUS
# ─────────────────────────────────────────────
async def menu_reports(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    keyboard = [
        [InlineKeyboardButton("📅 Today",      callback_data="rep_daily")],
        [InlineKeyboardButton("🗓️ This Week",  callback_data="rep_weekly")],
        [InlineKeyboardButton("📆 This Month", callback_data="rep_monthly")],
        [InlineKeyboardButton("⬅️ Back",       callback_data="back_main")],
    ]
    await query.edit_message_text("📊 *Select Report Period:*", parse_mode="Markdown",
                                   reply_markup=InlineKeyboardMarkup(keyboard))


async def menu_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    keyboard = [
        [InlineKeyboardButton("📅 Today (Excel)",      callback_data="exp_daily")],
        [InlineKeyboardButton("🗓️ This Week (Excel)",  callback_data="exp_weekly")],
        [InlineKeyboardButton("📆 This Month (Excel)", callback_data="exp_monthly")],
        [InlineKeyboardButton("⬅️ Back",               callback_data="back_main")],
    ]
    await query.edit_message_text("📥 *Export to Excel — Select Period:*", parse_mode="Markdown",
                                   reply_markup=InlineKeyboardMarkup(keyboard))


async def send_text_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query  = update.callback_query
    period = query.data.replace("rep_", "")
    await query.answer()
    text = generate_text_report(period)
    await query.edit_message_text(text, parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="menu_reports")]]))


async def send_excel_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query  = update.callback_query
    period = query.data.replace("exp_", "")
    await query.answer("Generating Excel file…")
    buf      = generate_excel_report(period)
    filename = f"logistics_{period}_{date.today().isoformat()}.xlsx"
    await context.bot.send_document(
        chat_id=query.message.chat_id,
        document=InputFile(buf, filename=filename),
        caption=f"📥 *{_period_label(period)}* — Excel Export",
        parse_mode="Markdown",
    )


# ─────────────────────────────────────────────
# LOG TRIP CONVERSATION
# Step 1: Choose Job
# Step 2: Choose Truck (from list or manual)
# Step 3: REMINDER — confirm truck plate
# Step 4: Enter Volume
# Step 5: Confirm & Save
# ─────────────────────────────────────────────
async def log_trip_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query   = update.callback_query
    user_id = query.from_user.id
    await query.answer()

    if not get_user_role(user_id):
        await query.edit_message_text("🚫 Access denied. Use your access link.")
        return ConversationHandler.END

    jobs = get_all_jobs(status_filter="active")
    if not jobs:
        await query.message.reply_text(
            "⚠️ *No active jobs available.*\n\nAsk an admin to add a job first.",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="back_main")]]),
        )
        return ConversationHandler.END

    buttons = [
        [InlineKeyboardButton(f"🏗️ {j['name']}", callback_data=f"tripjob_{j['id']}")]
        for j in jobs
    ]
    await query.message.reply_text(
        "📍 *Step 1 of 3 — Select Job Site:*",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(buttons),
    )
    return ASK_JOB


async def job_selected(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query  = update.callback_query
    job_id = int(query.data.replace("tripjob_", ""))
    await query.answer()

    job = get_job_by_id(job_id)
    context.user_data["job_name"]     = job["name"]
    context.user_data["job_location"] = job["location"] or ""

    # Show truck list
    trucks = get_all_trucks()
    if trucks:
        buttons = [
            [InlineKeyboardButton(f"🚛 {t['plate']}", callback_data=f"truckpick_{t['plate']}")]
            for t in trucks
        ]
        buttons.append([InlineKeyboardButton("✏️ Enter plate manually", callback_data="truckpick_manual")])
    else:
        buttons = [[InlineKeyboardButton("✏️ Enter plate manually", callback_data="truckpick_manual")]]

    await query.message.reply_text(
        f"🚛 *Step 2 of 3 — Select Truck Plate:*\n\n"
        f"Job: `{job['name']}`",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(buttons),
    )
    return ASK_PLATE


async def truck_selected_from_list(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    plate = query.data.replace("truckpick_", "")
    await query.answer()

    if plate == "manual":
        await query.message.reply_text(
            "✏️ *Enter the truck plate manually:*",
            parse_mode="Markdown",
        )
        return ASK_PLATE_MANUAL

    context.user_data["truck_plate"] = plate

    # ── REMINDER / CONFIRM ──────────────────
    keyboard = [
        [InlineKeyboardButton("✅ Yes, correct",      callback_data="plate_confirm_yes")],
        [InlineKeyboardButton("🔄 Choose different",  callback_data="plate_confirm_no")],
    ]
    await query.message.reply_text(
        f"⚠️ *Please confirm the truck plate:*\n\n"
        f"🚛 Plate:    `{plate}`\n"
        f"🏗️ Job Site: `{context.user_data['job_name']}`\n\n"
        f"Is this correct?",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return CONFIRM_TRIP


async def truck_entered_manually(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    plate = update.message.text.strip().upper()
    context.user_data["truck_plate"] = plate

    # ── REMINDER / CONFIRM ──────────────────
    keyboard = [
        [InlineKeyboardButton("✅ Yes, correct",      callback_data="plate_confirm_yes")],
        [InlineKeyboardButton("🔄 Choose different",  callback_data="plate_confirm_no")],
    ]
    await update.message.reply_text(
        f"⚠️ *Please confirm the truck plate:*\n\n"
        f"🚛 Plate:    `{plate}`\n"
        f"🏗️ Job Site: `{context.user_data['job_name']}`\n\n"
        f"Is this correct?",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return CONFIRM_TRIP


async def plate_confirmed(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    if query.data == "plate_confirm_no":
        # Go back to truck selection
        trucks = get_all_trucks()
        if trucks:
            buttons = [
                [InlineKeyboardButton(f"🚛 {t['plate']}", callback_data=f"truckpick_{t['plate']}")]
                for t in trucks
            ]
            buttons.append([InlineKeyboardButton("✏️ Enter plate manually", callback_data="truckpick_manual")])
        else:
            buttons = [[InlineKeyboardButton("✏️ Enter plate manually", callback_data="truckpick_manual")]]

        await query.message.reply_text(
            "🚛 *Select Truck Plate again:*",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(buttons),
        )
        return ASK_PLATE

    # Confirmed — ask for volume
    await query.message.reply_text(
        f"🧱 *Step 3 of 3 — Enter Volume poured (m³):*\n\n"
        f"🚛 Truck:    `{context.user_data['truck_plate']}`\n"
        f"🏗️ Job Site: `{context.user_data['job_name']}`",
        parse_mode="Markdown",
    )
    return ASK_VOLUME


async def ask_volume(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        volume = float(update.message.text.strip())
        if volume <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❌ Please enter a valid positive number (e.g. 8.5)")
        return ASK_VOLUME

    job_name    = context.user_data["job_name"]
    truck_plate = context.user_data["truck_plate"]
    save_trip(update.effective_user.id, job_name, truck_plate, volume)

    keyboard = [
        [InlineKeyboardButton("➕ Log Another Trip", callback_data="log_trip")],
        [InlineKeyboardButton("📊 View Reports",     callback_data="menu_reports")],
        [InlineKeyboardButton("🏠 Main Menu",        callback_data="back_main")],
    ]
    await update.message.reply_text(
        f"✅ *Trip Saved Successfully!*\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"📍 Job Site:  `{job_name}`\n"
        f"🚛 Truck:     `{truck_plate}`\n"
        f"🧱 Volume:    `{volume} m³`\n"
        f"🕐 Time:      `{datetime.now().strftime('%H:%M, %d %b %Y')}`",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    context.user_data.clear()
    return ConversationHandler.END


async def cancel_conv(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("❌ Action cancelled.")
    context.user_data.clear()
    return ConversationHandler.END


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    init_db()

    # Register the bot owner as admin on first run
    register_user(5613539602, "owner", "admin")

    app = Application.builder().token(BOT_TOKEN).build()

    # Add Job conversation
    add_job_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(add_job_start, pattern="^add_job$")],
        states={
            ASK_NEW_JOB_NAME:     [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_new_job_name)],
            ASK_NEW_JOB_LOCATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_new_job_location)],
        },
        fallbacks=[CommandHandler("cancel", cancel_conv)],
    )

    # Add Truck conversation
    add_truck_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(add_truck_start, pattern="^add_truck$")],
        states={
            ASK_NEW_TRUCK: [MessageHandler(filters.TEXT & ~filters.COMMAND, save_new_truck)],
        },
        fallbacks=[CommandHandler("cancel", cancel_conv)],
    )

    # Log Trip conversation
    log_trip_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(log_trip_start, pattern="^log_trip$")],
        states={
            ASK_JOB: [
                CallbackQueryHandler(job_selected, pattern="^tripjob_\\d+$"),
            ],
            ASK_PLATE: [
                CallbackQueryHandler(truck_selected_from_list, pattern="^truckpick_"),
            ],
            ASK_PLATE_MANUAL: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, truck_entered_manually),
            ],
            CONFIRM_TRIP: [
                CallbackQueryHandler(plate_confirmed, pattern="^plate_confirm_"),
            ],
            ASK_VOLUME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, ask_volume),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel_conv)],
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(add_job_conv)
    app.add_handler(add_truck_conv)
    app.add_handler(log_trip_conv)

    app.add_handler(CallbackQueryHandler(job_status,        pattern="^job_status$"))
    app.add_handler(CallbackQueryHandler(manage_trucks,     pattern="^manage_trucks$"))
    app.add_handler(CallbackQueryHandler(complete_job_menu, pattern="^complete_job$"))
    app.add_handler(CallbackQueryHandler(cancel_job_menu,   pattern="^cancel_job$"))
    app.add_handler(CallbackQueryHandler(do_complete_job,   pattern="^do_complete_\\d+$"))
    app.add_handler(CallbackQueryHandler(do_cancel_job,     pattern="^do_cancel_\\d+$"))
    app.add_handler(CallbackQueryHandler(back_main,         pattern="^back_main$"))
    app.add_handler(CallbackQueryHandler(menu_reports,      pattern="^menu_reports$"))
    app.add_handler(CallbackQueryHandler(menu_export,       pattern="^menu_export$"))
    app.add_handler(CallbackQueryHandler(send_text_report,  pattern="^rep_(daily|weekly|monthly)$"))
    app.add_handler(CallbackQueryHandler(send_excel_report, pattern="^exp_(daily|weekly|monthly)$"))

    logger.info("Bot is running…")
    app.run_polling()


if __name__ == "__main__":
    main()
